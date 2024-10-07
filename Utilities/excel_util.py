import openpyxl

class ExcelUtil:
    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_row_count(self):
        return self.sheet.max_row

    def get_cell_value(self, row, column):
        return self.sheet.cell(row=row, column=column).value