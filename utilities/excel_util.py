from openpyxl import load_workbook

class ExcelUtil:
    def __init__(self, EXCEL_PATH):
        self.excel_path = EXCEL_PATH
        self.workbook = load_workbook(filename=self.excel_path)
        self.sheet = self.workbook.active

    def get_row_count(self):
        """Get the total number of rows in the active sheet."""
        return self.sheet.max_row

    def get_cell_value(self, row, column):
        """Get the value of a cell in the specified row and column."""
        value = self.sheet.cell(row=row, column=column).value
        print(f"Reading value from row: {row}, column: {column} -> {value}")  # Debug statement
        if value:
            return str(value).strip()  # Convert to string and strip spaces
        return ""

    def write_cell_value(self, row, column, value):
        """
        Write a value to the specified cell in the active sheet.

        Parameters:
        - row: The row number of the cell.
        - column: The column number of the cell.
        - value: The value to write into the cell.
        """
        self.sheet.cell(row=row, column=column).value = value
        print(f"Writing value to row: {row}, column: {column} -> {value}")  # Debug statement
        self.workbook.save(self.excel_path)
        print(f"Workbook '{self.excel_path}' saved successfully.") # Confirmation message

